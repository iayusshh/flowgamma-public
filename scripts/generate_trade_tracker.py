"""
Simple signal tracker — Telegram calls vs Our System
Run: python3 scripts/generate_trade_tracker.py
Out: trade_tracker.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import os

# ── helpers ──────────────────────────────────────────────────────────────────
def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, color="FFFFFF", size=10):
    return Font(bold=bold, color=color, size=size, name="Consolas")

def border():
    s = Side(style="thin", color="1E2D3D")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, row, col, val, bg, fg="FFFFFF", bold=True, sz=9, wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg)
    c.font = Font(bold=bold, color=fg, size=sz, name="Consolas")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = border()
    return c

def dcell(ws, row, col, val=None, bg="0D1117", fg="BDC3C7", bold=False, fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = fill(bg)
    c.font = Font(bold=bold, color=fg, size=9, name="Consolas")
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = border()
    if fmt:
        c.number_format = fmt
    return c

SOURCES = ["Purvesh", "Trendmaster", "Chartians", "Other"]

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — SIGNAL LOG
# ─────────────────────────────────────────────────────────────────────────────
def build_log(wb):
    ws = wb.create_sheet("Signal Log")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "4FA3E0"
    ws.freeze_panes = "A3"

    # Column definitions: (header, width)
    cols = [
        ("#",          4),
        ("Date",      11),
        ("Stock",     16),
        ("Shared By", 14),
        ("Direction",  9),   # Buy / Sell
        ("Our System", 11),  # Bull / Bear / Neutral
        ("Aligns?",    9),   # auto formula
        ("Verdict",    9),   # Win / Loss / Pending
    ]

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    t = ws.cell(row=1, column=1, value="SIGNAL TRACKER")
    t.fill = fill("0A0F18")
    t.font = Font(bold=True, color="4FA3E0", size=14, name="Consolas")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for i, (hdr, w) in enumerate(cols, 1):
        hcell(ws, 2, i, hdr, "1A3A5C", sz=9)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 22

    # Data rows
    for r in range(3, 203):   # 200 rows
        bg = "10181F" if r % 2 == 0 else "0D1117"
        for i in range(1, len(cols) + 1):
            dcell(ws, r, i, bg=bg)

        # Auto row number
        ws.cell(row=r, column=1).value = r - 2

        # Aligns? formula: YES if (Direction=Buy AND System=Bull) OR (Sell AND Bear)
        # Col E = Direction, Col F = Our System, Col G = Aligns?
        ws.cell(row=r, column=7).value = (
            f'=IF(OR(AND(E{r}="Buy",F{r}="Bull"),AND(E{r}="Sell",F{r}="Bear")),'
            f'"YES",IF(OR(F{r}="",E{r}=""),"","NO"))'
        )
        ws.cell(row=r, column=7).fill  = fill(bg)
        ws.cell(row=r, column=7).font  = Font(color="BDC3C7", size=9, name="Consolas")
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=7).border = border()

        ws.cell(row=r, column=2).number_format = "DD-MMM-YY"

    # ── Dropdowns ──
    def dv(formula, sqref):
        v = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(v)
        v.sqref = sqref

    dv(f'"{",".join(SOURCES)}"', "D3:D202")
    dv('"Buy,Sell"',              "E3:E202")
    dv('"Bull,Bear,Neutral"',     "F3:F202")
    dv('"Win,Loss,Pending"',      "H3:H202")

    # ── Conditional formatting ──
    # Aligns YES = green, NO = red
    ws.conditional_formatting.add("G3:G202", CellIsRule(
        operator="equal", formula=['"YES"'],
        fill=fill("0F2A1A"), font=Font(color="27AE60", bold=True, name="Consolas", size=9)))
    ws.conditional_formatting.add("G3:G202", CellIsRule(
        operator="equal", formula=['"NO"'],
        fill=fill("2A0F0F"), font=Font(color="E74C3C", bold=True, name="Consolas", size=9)))

    # Verdict WIN = green, LOSS = red, PENDING = amber
    ws.conditional_formatting.add("H3:H202", CellIsRule(
        operator="equal", formula=['"Win"'],
        fill=fill("0F2A1A"), font=Font(color="27AE60", bold=True, name="Consolas", size=9)))
    ws.conditional_formatting.add("H3:H202", CellIsRule(
        operator="equal", formula=['"Loss"'],
        fill=fill("2A0F0F"), font=Font(color="E74C3C", bold=True, name="Consolas", size=9)))
    ws.conditional_formatting.add("H3:H202", CellIsRule(
        operator="equal", formula=['"Pending"'],
        fill=fill("2A1A0A"), font=Font(color="F39C12", bold=True, name="Consolas", size=9)))

    # Direction Buy = green tint, Sell = red tint
    ws.conditional_formatting.add("E3:E202", CellIsRule(
        operator="equal", formula=['"Buy"'],
        fill=fill("0F2A1A"), font=Font(color="27AE60", bold=True, name="Consolas", size=9)))
    ws.conditional_formatting.add("E3:E202", CellIsRule(
        operator="equal", formula=['"Sell"'],
        fill=fill("2A0F0F"), font=Font(color="E74C3C", bold=True, name="Consolas", size=9)))

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — ACCURACY DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────
def build_accuracy(wb):
    ws = wb.create_sheet("Accuracy")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "27AE60"

    for col, w in [("A",22),("B",12),("C",12),("D",14),("E",12),("F",14),("G",22)]:
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1, value="ACCURACY TRACKER")
    t.fill = fill("0A0F18")
    t.font = Font(bold=True, color="4FA3E0", size=14, name="Consolas")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    def sec(row, text, col_end="G"):
        ws.merge_cells(f"A{row}:{col_end}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill("1A3A5C")
        c.font = Font(bold=True, color="4FA3E0", size=9, name="Consolas")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = border()
        ws.row_dimensions[row].height = 18

    def row_hdr(row, labels):
        for i, lbl in enumerate(labels, 1):
            hcell(ws, row, i, lbl, "1E2D3D", sz=8, wrap=True)
        ws.row_dimensions[row].height = 26

    def stat_row(row, label, *formulas_and_fmts, bg="0D1117"):
        dcell(ws, row, 1, label, bg=bg, fg="BDC3C7", align="left")
        for i, item in enumerate(formulas_and_fmts, 2):
            if isinstance(item, tuple):
                val, fmt = item
            else:
                val, fmt = item, None
            c = dcell(ws, row, i, val, bg=bg, fg="FFFFFF", bold=True)
            if fmt:
                c.number_format = fmt

    # ── SECTION 1: Per-source breakdown ──
    r = 2
    sec(r, "BY SOURCE — Signal Alignment (Does their call match our system?)")
    r += 1
    row_hdr(r, ["Source", "Total\nSignals", "Aligned\n(#)", "Aligned\n(%)",
                "Wins\n(#)", "Win %", "Not Aligned\n→ Win %"])
    r += 1

    for src in SOURCES:
        bg = "10181F" if r % 2 == 0 else "0D1117"
        dcell(ws, r, 1, src, bg=bg, fg="BDC3C7", align="left")

        total   = f'=COUNTIF(\'Signal Log\'!D3:D202,"{src}")'
        aligned = f'=COUNTIFS(\'Signal Log\'!D3:D202,"{src}",\'Signal Log\'!G3:G202,"YES")'
        al_pct  = f"=IFERROR({aligned}/B{r},0)"
        wins    = f'=COUNTIFS(\'Signal Log\'!D3:D202,"{src}",\'Signal Log\'!H3:H202,"Win")'
        win_pct = f"=IFERROR({wins}/B{r},0)"
        # Wins where NOT aligned
        na_wins = f'=COUNTIFS(\'Signal Log\'!D3:D202,"{src}",\'Signal Log\'!G3:G202,"NO",\'Signal Log\'!H3:H202,"Win")'
        na_tot  = f'=COUNTIFS(\'Signal Log\'!D3:D202,"{src}",\'Signal Log\'!G3:G202,"NO")'
        na_pct  = f"=IFERROR({na_wins}/{na_tot},0)"

        for col, (val, fmt) in enumerate([
            (total, "0"), (aligned, "0"), (al_pct, "0.0%"),
            (wins, "0"), (win_pct, "0.0%"), (na_pct, "0.0%")
        ], 2):
            c = dcell(ws, r, col, val, bg=bg, fg="FFFFFF", bold=False, fmt=fmt)

        # Color the aligned % cell
        ws.conditional_formatting.add(f"D{r}:D{r}", CellIsRule(
            operator="greaterThan", formula=["0.6"],
            fill=fill("0F2A1A"), font=Font(color="27AE60", bold=True, name="Consolas", size=9)))
        ws.conditional_formatting.add(f"D{r}:D{r}", CellIsRule(
            operator="lessThan", formula=["0.4"],
            fill=fill("2A0F0F"), font=Font(color="E74C3C", bold=True, name="Consolas", size=9)))
        r += 1

    # TOTAL row
    total_r = r
    dcell(ws, r, 1, "TOTAL", bg="1A3A5C", fg="4FA3E0", bold=True, align="left")
    for col, (val, fmt) in enumerate([
        (f"=SUM(B{r-len(SOURCES)}:B{r-1})", "0"),
        (f"=SUM(C{r-len(SOURCES)}:C{r-1})", "0"),
        (f"=IFERROR(C{r}/B{r},0)", "0.0%"),
        (f"=SUM(E{r-len(SOURCES)}:E{r-1})", "0"),
        (f"=IFERROR(E{r}/B{r},0)", "0.0%"),
        ("", None),
    ], 2):
        dcell(ws, r, col, val, bg="1A3A5C", fg="FFFFFF", bold=True, fmt=fmt)
    r += 2

    # ── SECTION 2: Overall at-a-glance ──
    sec(r, "OVERALL SUMMARY")
    r += 1

    stats = [
        ("Total Signals Logged",
         f'=COUNTA(\'Signal Log\'!C3:C202)', "0"),
        ("Total Aligned with System",
         f'=COUNTIF(\'Signal Log\'!G3:G202,"YES")', "0"),
        ("System Alignment Rate",
         f'=IFERROR(COUNTIF(\'Signal Log\'!G3:G202,"YES")/COUNTA(\'Signal Log\'!C3:C202),0)', "0.0%"),
        ("Total Wins (all sources)",
         f'=COUNTIF(\'Signal Log\'!H3:H202,"Win")', "0"),
        ("Total Losses",
         f'=COUNTIF(\'Signal Log\'!H3:H202,"Loss")', "0"),
        ("Overall Win Rate",
         f'=IFERROR(COUNTIF(\'Signal Log\'!H3:H202,"Win")/COUNTIFS(\'Signal Log\'!H3:H202,"<>",\'Signal Log\'!H3:H202,"<>Pending"),0)', "0.0%"),
        ("Win Rate when Aligned",
         f'=IFERROR(COUNTIFS(\'Signal Log\'!G3:G202,"YES",\'Signal Log\'!H3:H202,"Win")/COUNTIFS(\'Signal Log\'!G3:G202,"YES",\'Signal Log\'!H3:H202,"<>",\'Signal Log\'!H3:H202,"<>Pending"),0)', "0.0%"),
        ("Win Rate when NOT Aligned",
         f'=IFERROR(COUNTIFS(\'Signal Log\'!G3:G202,"NO",\'Signal Log\'!H3:H202,"Win")/COUNTIFS(\'Signal Log\'!G3:G202,"NO",\'Signal Log\'!H3:H202,"<>",\'Signal Log\'!H3:H202,"<>Pending"),0)', "0.0%"),
    ]

    row_hdr(r, ["Metric", "Value", "", "", "", "", ""])
    ws.merge_cells(f"C{r}:G{r}")
    r += 1

    for i, (label, formula, fmt) in enumerate(stats):
        bg = "10181F" if i % 2 == 0 else "0D1117"
        dcell(ws, r, 1, label, bg=bg, fg="BDC3C7", align="left")
        c = dcell(ws, r, 2, formula, bg=bg, fg="FFFFFF", bold=True, fmt=fmt)
        ws.merge_cells(f"C{r}:G{r}")
        ws.cell(row=r, column=3).fill = fill(bg)
        ws.cell(row=r, column=3).border = border()
        r += 1

    return ws


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)
    build_log(wb)
    build_accuracy(wb)

    out = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "trade_tracker.xlsx"))
    wb.save(out)
    print(f"Saved → {out}")

if __name__ == "__main__":
    main()
